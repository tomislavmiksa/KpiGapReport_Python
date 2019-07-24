# USED FOR DB ACCESS
import pyodbc
import sqlalchemy
# PYTHON DATA ANALYSIS LIBRARY
import pandas as pd
# WRITING TO EXCEL FILE
import xlsxwriter
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import Rule, IconSet, FormatObject

# local scripts and libs
import sys
import define_classes as kc
import script_calc_kpi as ck

# INPUT INFO
op = kc.InputInfo(server         = "blndb11",
                  database       = "DE_BM_Voice_1905",
                  kpi_report     = ['NEW_KPI_OPERATOR_1', 'NEW_KPI_OPERATOR_2', 'NEW_KPI_OPERATOR_3'],
                  operator       = ['Telekom', 'Vodafone', 'Telefonica'],
                  colors         = ['E20074','DA291C','014268'])

vdb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( op.server, op.database ) )
vcur = vdb.cursor()
vdf_target = pd.read_sql("SELECT * FROM VDF_TARGET_KPI", vdb)

i = 0
for view in op.kpi_report:
    test = pd.read_sql("SELECT * FROM " + view, vdb)
    test = test[ck.column_order()]
    test.insert(loc=0, column='index', value=0)
    test.insert(loc = 0, column = 'Operator', value = op.operator[i])
    test.insert(loc=0, column='Operator_Order', value=i)
    op.kpi_report_sql.append(test)
    del test
    i = i + 1

writer = xlsxwriter.Workbook("kpi_report.xlsx")
# WRITE TO G_LEVEL_0
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g0',
               writer = writer,
               sheet = 'OVERALL',
               g_lev_1='',
               g_lev_2='')

# WRITE TO G_LEVEL_2
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g2',
               writer = writer,
               sheet = 'ALL PER RANKING MODULE',
               g_lev_1='',
               g_lev_2='')

# WRITE TO VDF_G_LEVEL_2
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g2',
               writer = writer,
               sheet = 'VDF PER RANKING MODULE',
               g_lev_1='',
               g_lev_2='',
               ind = [1])

# WRITE TO VDF_DTK_G_LEVEL_2
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g2',
               writer = writer,
               sheet = 'GAP TO TELEKOM RANKING',
               g_lev_1='',
               g_lev_2='',
               ind = [0,1])

# WRITE TO G_LEVEL_2 WITH VENDOR
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g2v',
#               writer = writer,
#               sheet = 'G_Level_2_Vendor',
#               g_lev_1='',
#               g_lev_2='')

# WRITE TO VDF_G_LEVEL_2
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g2v',
#               writer = writer,
#               sheet = 'VDF_G_Level_2_Vendor',
#               g_lev_1='',
#               g_lev_2='',
#               ind = [1])

# WRITE TO VDF_DTK_G_LEVEL_2
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g2v',
#               writer = writer,
#               sheet = 'VDF_DTK_G_LEVEL_2_Vendor',
#               g_lev_1='',
#               g_lev_2='',
#               ind = [0,1])

# WRITE TO G_LEVEL_3
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g3',
#               writer = writer,
#               sheet = 'G_Level_3',
#               g_lev_1='',
#               g_lev_2='')

# WRITE TO G_LEVEL_3 WITH VENDOR
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g3v',
#               writer = writer,
#               sheet = 'G_Level_3_Vendor',
#               g_lev_1='',
#               g_lev_2='')

# WRITE TO VDF_G_LEVEL_3
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g3v',
               writer = writer,
               sheet = 'VDF PER MODULE PER VENDOR',
               g_lev_1='',
               g_lev_2='',
               ind = [1])

# WRITE TO VDF_DTK_G_LEVEL_3
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g3',
               writer = writer,
               sheet = 'GAP TO TELEKOM PER MODULE',
               g_lev_1='',
               g_lev_2='',
               ind = [0,1])

## WRITE TO G_LEVEL_4 CITY
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g4',
#               writer = writer,
#               sheet = 'CITY',
#               g_lev_1='',
#               g_lev_2='City')
#
## WRITE TO G_LEVEL_4 ROADS
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g4',
#               writer = writer,
#               sheet = 'CONN ROADS',
#               g_lev_1='',
#               g_lev_2='Connecting Roads')
#
## WRITE TO G_LEVEL_4 TRAIN
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g4',
#               writer = writer,
#               sheet = 'TRAIN',
#               g_lev_1='',
#               g_lev_2='Train Route')
#
## WRITE TO G_LEVEL_5
#ck.write_excel(reports=op.kpi_report_sql,
#               group = 'g5',
#               writer = writer,
#               sheet = 'HOTSPOTS',
#               g_lev_1='',
#               g_lev_2='')

# WRITE TO G_LEVEL_4 CITY
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g4',
               writer = writer,
               sheet = 'CITY',
               g_lev_1='',
               g_lev_2='City',
               ind = [1])

# WRITE TO G_LEVEL_4 ROADS
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g4',
               writer = writer,
               sheet = 'CONN ROADS',
               g_lev_1='',
               g_lev_2='Connecting Roads',
               ind = [1])

# WRITE TO G_LEVEL_4 TRAIN
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g4',
               writer = writer,
               sheet = 'TRAIN',
               g_lev_1='',
               g_lev_2='Train Route',
               ind = [1])

# WRITE TO G_LEVEL_5
ck.write_excel(reports=op.kpi_report_sql,
               group = 'g5',
               writer = writer,
               sheet = 'HOTSPOTS',
               g_lev_1='',
               g_lev_2='',
               ind = [1])

writer.close()

# conditional formating
workbook = xl.load_workbook(r"kpi_report.xlsx")
for sheet in workbook.worksheets:
    for cell in sheet[1]:
        try:
            # OPERATOR COLORING PART
            font = Font(name='Calibri',
                        size = 13,
                        bold = True,
                        italic = False,
                        vertAlign = None,
                        underline = 'none',
                        strike = False,
                        color = 'FFFFFF')
            num = int(sheet.cell(row=1, column=cell.column).value)
            col = op.colors[num]
            sheet.cell(row=2,column=cell.column).fill = xl.styles.PatternFill(fgColor=col, fill_type = "solid")
            sheet.cell(row=2,column=cell.column).font = font

            # CONDITIONAL FORMATING PART
            glev1 = sheet.cell(row=6,column=cell.column).value
            glev2 = sheet.cell(row=7,column=cell.column).value
            for index,row in vdf_target.iterrows():
                if glev1 == "Drive" and glev2 == "City":
                    thr = row.DC
                elif glev1 == "Drive" and glev2 == "Connecting Roads":
                    thr = row.DR
                elif glev1 == "Walk" and glev2 == "City":
                    thr = row.WC
                elif glev1 == "Walk" and glev2 == "Train Route":
                    thr = row.WT

                if row.Reverse > 0:
                    rev_flag = True
                    cond = [FormatObject(type='num', val=thr),
                            FormatObject(type='num', val=thr),
                            FormatObject(type='num', val=thr)]
                else:
                    rev_flag = None
                    cond = [FormatObject(type='num', val=row.MIN),
                            FormatObject(type='num', val=thr),
                            FormatObject(type='num', val=thr)]

                rule = Rule(type='iconSet',
                            iconSet=IconSet(iconSet='3TrafficLights1',
                                            cfvo=cond,
                                            showValue=None,
                                            percent=None,
                                            reverse=rev_flag))

                ref_cell = ck.colnum_string(cell.column) + str(int(row.ROW))
                # sheet.conditional_formatting.add(ref_cell, rule)

                # sheet.cell(ref_cell).style.font.bold = True
                # sheet.cell(ref_cell).style.font.size = 14
        except:
            pass

    # glev_flag = sheet.cell(row=4, column=2).value
    # sheet.delete_rows(idx=13, amount=1)
    # #sheet.row_dimensions(13).hidden = True

    #if glev_flag:
    #    if glev_flag.find('v') == -1:
    #        sheet.delete_rows(idx=12, amount=1)

    #sheet.delete_rows(idx=3, amount=3)
    #sheet.delete_rows(idx=1, amount=1)

workbook.save("kpi_report.xlsx")

