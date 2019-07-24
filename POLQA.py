# USED FOR DB ACCESS
import pyodbc
# PYTHON DATA ANALYSIS LIBRARY
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt

# WRITING TO EXCEL FILE
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import Rule, IconSet, FormatObject

# FUNCTION - WRITE TO EXCEL TO AVOID CODE REPRETITIONS
########################################################################################################################
def write_pd(df_tmp,book,sheet_name):
    sheet_tmp = book.create_sheet(sheet_name)
    j = 1
    for v in list(df_tmp.columns.values):
        try:
            sheet_tmp.cell(row=1, column=j).value = v
            sheet_tmp.title = sheet_name
        except:
            print('Header Printing Failed')
            pass
        j += 1
    i = 1
    for value in list(df_tmp):
        j = 2
        for v in list(df_tmp[value]):
            try:
                sheet_tmp.cell(row=j, column=i).value = v
            except:
                print('Writting in row', i, ' and collumn ', j, 'FAILED!')
                pass
            j += 1
        i += 1


srv = "blndb11"
db  = "DE_BM_Voice_1905"
tb =  "vSpeechCDR2018_Operator1"

vdb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( srv, db ) )
vcur = vdb.cursor()
cdr_voice = pd.read_sql('select * from ' + tb, vdb)

sns.set(style="whitegrid")
data = cdr_voice.pivot_table(cdr_voice,index="LQ",
                             columns= "BW",
                             values = 'ReceiveDelay',
                             aggfunc=lambda x: len(x.avg()))
#ax = sns.heatmap(data)
print(data)
# plt.show()

# WRITE KPI REPORT AS IN DATABASE
#book = Workbook()
#write_pd(cdr_voice, book, 'SPEECH_CDR')
#book.save('tm_speech.xlsx')
