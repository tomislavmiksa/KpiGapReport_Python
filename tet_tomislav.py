# USED FOR DB ACCESS
import pyodbc
import sqlalchemy
# PYTHON DATA ANALYSIS LIBRARY
import pandas as pd
# WRITING TO EXCEL FILE
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import Rule, IconSet, FormatObject

# CLASES
########################################################################################################################
class OperatorCDRs:
    def __init__(self,server,voice_db,data_db,voice_table,speech_table,data_table, output_table):
        self.srv = server
        self.v_db = voice_db
        self.d_db = data_db
        self.v_cdr = voice_table
        self.s_cdr = speech_table
        self.d_cdr = data_table
        self.table_name = output_table

class GeoStuff:
    def __init__(self, g1, g2):
        self.g1 = g1
        self.g2 = g2
        self.g2v = []
        self.g3 = []
        self.g3v = []
        self.g4 = []
        self.g5 = []
        self.tn = []

    def updateLocations(self, cdr):
        _tmp = cdr
        _tmp['G_Level_2_Vendor'] = _tmp['G_Level_2'] + ';' + _tmp['Vendor']
        _tmp['G_Level_3_Vendor'] = _tmp['G_Level_3'] + ';' + _tmp['Vendor']
        _tmp['G_Level_34']       = _tmp['G_Level_3'] + ';' + _tmp['G_Level_4']
        _tmp['G_Level_45']       = _tmp['G_Level_4'] + ';' + _tmp['G_Level_5']
        _tmp['TrainRoute']       = _tmp['G_Level_4'] + ';' + _tmp['Train'] + ';' + cdr_voice['Wagon']

        _cdr = cdr[ (cdr.G_Level_1 == self.g1) & (cdr.G_Level_2 == self.g2) ]

        self.g2v.extend(_cdr.G_Level_2_Vendor)
        self.g2v = list(set(self.g2v))
        self.g2v.sort()

        self.g3.extend(_cdr.G_Level_3)
        self.g3 = list(set(self.g3))
        self.g3.sort()

        self.g3v.extend(_cdr.G_Level_3_Vendor)
        self.g3v = list(set(self.g3v))
        self.g3v.sort()

        self.g4.extend(_cdr.G_Level_34)
        self.g4 = list(set(self.g4))
        self.g4.sort()

        if (self.g1 == "Walk") & (self.g2 == "City"):
            self.g5.extend(_cdr.G_Level_45)
            self.g5 = list(set(self.g5))
            self.g5.sort()

        if (self.g2 == "Train Route"):
            self.tn.extend(_cdr.TrainRoute)
            self.tn = list(set(self.tn))
            self.tn.sort()

    def printLocations(self):
        print("\n")
        print("G_Level_1               :", self.g1)
        print("G_Level_2               :", self.g2)
        print("G_Level_2;Vendor        :", self.g2v)
        print("G_Level_3               :", self.g3)
        print("G_Level_3;Vendor        :", self.g3v)
        print("G_Level_4               :", self.g4)
        print("G_Level_5               :", self.g5)
        print("Train; Wagon            :", self.tn)

class Voice:
    def __init__(self,records,g1='',g2='',g3='',g4='',g5='',vend = '',reg ='',t_name = '',w_name = ''):
        self.g1 = g1
        self.g2 = g2
        self.g3 = g3
        self.g4 = g4
        self.g5 = g5
        self.vendor = vend
        self.region = reg
        self.train = t_name
        self.wagon = w_name
        _records = records
        # FILTERING CORRECT DATA
        if g1 != '':
            _records_filter = _records[_records["G_Level_1"] == g1]
            _records = _records_filter
        if g2 != '':
            _records_filter = _records[_records["G_Level_2"] == g2]
            _records = _records_filter
        if g3 != '':
            _records_filter = _records[_records["G_Level_3"] == g3]
            _records = _records_filter
        if g4 != '':
            _records_filter = _records[_records["G_Level_4"] == g4]
            _records = _records_filter
        if g5 != '':
            _records_filter = _records[_records["G_Level_5"] == g5]
            _records = _records_filter
        if vend != '':
            _records_filter = _records[_records["Vendor"] == vend]
            _records = _records_filter
        if reg != '':
            _records_filter = _records[_records["Region"] == reg]
            _records = _records_filter
        if t_name != '':
            _records_filter = _records[_records["Fleet"] == t_name]
            _records = _records_filter
        if w_name != '':
            _records_filter = _records[_records["WagonNumber"] == w_name]
            _records = _records_filter

        _classic  = _records[_records.Session_Type == 'CALL']
        _whatsapp = _records[_records.Session_Type == 'WhatsApp CALL']

        # START AND END TIME
        ################################################################################################################
        _perc = [.10, .20, .50, .80, .90]
        _tmp1 = _classic["Call_Start_Time"].describe(percentiles=_perc)
        self.start_time = _tmp1.loc['min']
        self.end_time   = _tmp1.loc['max']

        # CLASSIC CALL STATISTICS
        ################################################################################################################
        self.classic_attempts = len(_classic)
        self.classic_completed = len(_classic[_classic.Call_Status == 'Completed'])
        self.classic_failed = len(_classic[_classic.Call_Status == 'Failed'])
        self.classic_dropped = len(_classic[_classic.Call_Status == 'Dropped'])
        # Call Setup Success Ratio
        if self.classic_attempts > 0:
            self.classic_cssr = (self.classic_attempts - self.classic_failed)/self.classic_attempts
        else:
            self.classic_cssr = 0
        # Droped Call Ratio
        if (self.classic_attempts-self.classic_failed) > 0:
            self.classic_dcr = self.classic_dropped / (self.classic_attempts-self.classic_failed)
        else:
            self.classic_dcr = 0
        # Call Success Ratio
        if self.classic_attempts > 0:
            self.classic_ccr = self.classic_completed/self.classic_attempts
        else:
            self.classic_ccr = 0

        # CALL SETUP TIME
        ################################################################################################################
        _classic1 = _classic[_classic.Call_Status == 'Completed']
        _tmp = _classic1["CST(Dial->ConnAck)"].describe(percentiles=_perc)

        self.classic_min_cst = _tmp.loc['min']
        self.classic_avg_cst = _tmp.loc['mean']
        self.classic_max_cst = _tmp.loc['max']
        self.classic_p10_cst = _tmp.loc['10%']
        self.classic_p50_cst = _tmp.loc['50%']
        self.classic_p90_cst = _tmp.loc['90%']

        _classic.columns = _classic.columns.str.replace('->', '')
        _classic.columns = _classic.columns.str.replace('(', '')
        _classic.columns = _classic.columns.str.replace(')', '')

        self.classic_poorCST = len(_classic[_classic.CSTDialConnAck > 15.0])
        self.classic_goodCST = len(_classic[_classic.CSTDialConnAck <= 15.0])
        if (self.classic_poorCST + self.classic_goodCST) > 0:
            self.classic_poorCSTratio = self.classic_poorCST / (self.classic_poorCST + self.classic_goodCST)
        else:
            self.classic_poorCSTratio = 0

        # VoLTE Call Mode
        ################################################################################################################
        self.volte_start = len(_classic[_classic.L1_callMode_A == 'VoLTE']) + len(_classic[_classic.L1_callMode_B == 'VoLTE'])
        if self.classic_attempts > 0:
            self.volte_ratio = 0.5 * self.volte_start / self.classic_attempts
        else:
            self.volte_ratio = 0
        self.volte_end = len(_classic[(_classic.L2_callMode_A == 'VoLTE') & (_classic.L2_callMode_B == 'VoLTE')])
        if self.classic_attempts > 0:
            self.volte_end_ratio = self.volte_end / self.classic_attempts
        else:
            self.volte_end_ratio = 0
        self.csfb = len(_classic[_classic.L1_callMode_A == 'CSFB']) + len(_classic[_classic.L1_callMode_B == 'CSFB'])
        if self.classic_attempts > 0:
            self.csfb_ratio = 0.5 * self.csfb / self.classic_attempts
        else:
            self.csfb_ratio = 0

        # WHATSAPP CALL STATISTICS
        ################################################################################################################
        self.whatsapp_attempts = len(_whatsapp)
        self.whatsapp_completed = len(_whatsapp[_whatsapp.Call_Status == 'Completed'])
        self.whatsapp_failed = len(_whatsapp[_whatsapp.Call_Status == 'Failed'])
        self.whatsapp_dropped = len(_whatsapp[_whatsapp.Call_Status == 'Dropped'])
        # Call Setup Success Ratio
        if self.whatsapp_attempts > 0:
            self.whatsapp_cssr = (self.whatsapp_attempts - self.whatsapp_failed)/self.whatsapp_attempts
        else:
            self.whatsapp_cssr = 0
        # Droped Call Ratio
        if (self.whatsapp_attempts-self.whatsapp_failed) > 0:
            self.whatsapp_dcr = self.whatsapp_dropped / (self.whatsapp_attempts-self.whatsapp_failed)
        else:
            self.whatsapp_dcr = 0
        # Call Success Ratio
        if self.whatsapp_attempts > 0:
            self.whatsapp_ccr = self.whatsapp_completed/self.whatsapp_attempts
        else:
            self.whatsapp_ccr = 0

        # WHATSAPP CALL SETUP TIME
        ################################################################################################################
        _perc = [.10, .20, .50, .80, .90]
        _whatsapp1 = _whatsapp[_whatsapp.Call_Status == 'Completed']
        _tmp = _whatsapp1["CST(Dial->ConnAck)"].describe(percentiles=_perc)

        self.whatsapp_min_cst = _tmp.loc['min']
        self.whatsapp_avg_cst = _tmp.loc['mean']
        self.whatsapp_max_cst = _tmp.loc['max']
        self.whatsapp_p10_cst = _tmp.loc['10%']
        self.whatsapp_p50_cst = _tmp.loc['50%']
        self.whatsapp_p90_cst = _tmp.loc['90%']

        _whatsapp.columns = _whatsapp.columns.str.replace('->', '')
        _whatsapp.columns = _whatsapp.columns.str.replace('(', '')
        _whatsapp.columns = _whatsapp.columns.str.replace(')', '')

        self.whatsapp_poorCST = len(_whatsapp[_whatsapp.CSTDialConnAck > 15.0])
        self.whatsapp_goodCST = len(_whatsapp[_whatsapp.CSTDialConnAck <= 15.0])
        if (self.whatsapp_poorCST + self.whatsapp_goodCST) > 0:
            self.whatsapp_poorCSTratio = self.whatsapp_poorCST / (self.whatsapp_poorCST + self.whatsapp_goodCST)
        else:
            self.whatsapp_poorCSTratio = 0

class Speech:
    def __init__(self,records,g1='',g2='',g3='',g4='',g5='',vend = "",reg = "",t_name = '',w_name = ''):
        self.g1 = g1
        self.g2 = g2
        self.g3 = g3
        self.g4 = g4
        self.g5 = g5
        self.vendor = vend
        self.region = reg
        self.train = t_name
        self.wagon = w_name
        _records = records
        # FILTERING CORRECT DATA
        if g1 != '':
            _records_filter = _records[_records["G_Level_1"] == g1]
            _records = _records_filter
        if g2 != '':
            _records_filter = _records[_records["G_Level_2"] == g2]
            _records = _records_filter
        if g3 != '':
            _records_filter = _records[_records["G_Level_3"] == g3]
            _records = _records_filter
        if g4 != '':
            _records_filter = _records[_records["G_Level_4"] == g4]
            _records = _records_filter
        if g5 != '':
            _records_filter = _records[_records["G_Level_5"] == g5]
            _records = _records_filter
        if vend != '':
            _records_filter = _records[_records["Vendor"] == vend]
            _records = _records_filter
        if reg != '':
            _records_filter = _records[_records["Region"] == reg]
            _records = _records_filter
        if t_name != '':
            _records_filter = _records[_records["Fleet"] == t_name]
            _records = _records_filter
        if w_name != '':
            _records_filter = _records[_records["WagonNumber"] == w_name]
            _records = _records_filter

        _classic  = _records[_records.Session_Type == 'CALL']
        _whatsapp = _records[_records.Session_Type == 'WhatsApp CALL']


        # POLQA SAMPLES COUNT
        self.classic_attempts = len(_classic[_classic.LQ > 0])
        self.classic_poorPMOS = len(_classic[_classic.LQ <= 1.6])
        if self.classic_attempts > 0:
            self.classic_poorRMOS = self.classic_poorPMOS/self.classic_attempts
        else:
            self.classic_poorRMOS = 0
        _perc = [.10, .20, .50, .80, .90]
        _tmp = _classic["LQ"].describe(percentiles=_perc)
        self.classic_min_mos = _tmp.loc['min']
        self.classic_avg_mos = _tmp.loc['mean']
        self.classic_max_mos = _tmp.loc['max']
        self.classic_p10_mos = _tmp.loc['10%']
        self.classic_p50_mos = _tmp.loc['50%']
        self.classic_p90_mos = _tmp.loc['90%']

        _tmpv = _classic["L1_VoLTE_Call_Mode_LQ"].describe(percentiles=_perc)

        self.classic_min_vmos = _tmpv.loc['min']
        self.classic_avg_vmos = _tmpv.loc['mean']
        self.classic_max_vmos = _tmpv.loc['max']
        self.classic_p10_vmos = _tmpv.loc['10%']
        self.classic_p50_vmos = _tmpv.loc['50%']
        self.classic_p90_vmos = _tmpv.loc['90%']

       # WHATSAPP MOS STATISTICS
        self.whatsapp_attempts = len(_whatsapp[_whatsapp.LQ > 0])
        self.whatsapp_poorPMOS = len(_whatsapp[_whatsapp.LQ <= 1.6])
        if self.whatsapp_attempts > 0:
            self.whatsapp_poorRMOS = self.whatsapp_poorPMOS/self.whatsapp_attempts
        else:
            self.whatsapp_poorRMOS = 0
        _perc = [.10, .20, .50, .80, .90]
        _tmp = _whatsapp["LQ"].describe(percentiles=_perc)
        self.whatsapp_min_mos = _tmp.loc['min']
        self.whatsapp_avg_mos = _tmp.loc['mean']
        self.whatsapp_max_mos = _tmp.loc['max']
        self.whatsapp_p10_mos = _tmp.loc['10%']
        self.whatsapp_p50_mos = _tmp.loc['50%']
        self.whatsapp_p90_mos = _tmp.loc['90%']

class Data:
    def __init__(self,records,g1='',g2='',g3='',g4='',g5='',vend = "",reg = "",t_name = '',w_name = ''):
        self.g1 = g1
        self.g2 = g2
        self.g3 = g3
        self.g4 = g4
        self.g5 = g5
        self.vendor = vend
        self.region = reg
        self.train = t_name
        self.wagon = w_name
        _records = records
        # FILTERING CORRECT DATA
        if g1 != '':
            _records_filter = _records[_records["G_Level_1"] == g1]
            _records = _records_filter
        if g2 != '':
            _records_filter = _records[_records["G_Level_2"] == g2]
            _records = _records_filter
        if g3 != '':
            _records_filter = _records[_records["G_Level_3"] == g3]
            _records = _records_filter
        if g4 != '':
            _records_filter = _records[_records["G_Level_4"] == g4]
            _records = _records_filter
        if g5 != '':
            _records_filter = _records[_records["G_Level_5"] == g5]
            _records = _records_filter
        if vend != '':
            _records_filter = _records[_records["Vendor"] == vend]
            _records = _records_filter
        if reg != '':
            _records_filter = _records[_records["Region"] == reg]
            _records = _records_filter
        if t_name != '':
            _records_filter = _records[_records["Train_Name"] == t_name]
            _records = _records_filter
        if w_name != '':
            _records_filter = _records[_records["Wagon_Number"] == w_name]
            _records = _records_filter

        try:
            _perc = [.10, .20, .50, .80, .90]
            _tmp1 = _records["Test_Start_Time"].describe(percentiles=_perc)
            self.start_time = _tmp1.loc['min']
            self.end_time   = _tmp1.loc['max']
        except:
            self.start_time = None
            self.end_time   = None

        # HTTP TRANSFER RAT STATISTICS
        _http_tran = _records[_records.Type_of_Test == 'httpTransfer']
        if len(_records) > 0:
            self.lte_share = _records.PCell_RAT.str.count("LTE").sum() / len(_records)
        else:
            self.lte_share = 0

            # SMALL FILE DOWNLOAD
        _http_fdfs_dl = _records[(_records.Type_of_Test == 'httpTransfer') & (_records.Test_Name == 'FDFS http DL ST')]
        self.fdfs_dl_attempts = len(_http_fdfs_dl)
        self.fdfs_dl_failed   = len(_http_fdfs_dl[_http_fdfs_dl.Test_Result == 'Failed'])
        self.fdfs_dl_cutoff   = len(_http_fdfs_dl[_http_fdfs_dl.Test_Result == 'Cutoff'])
        self.fdfs_dl_succes   = len(_http_fdfs_dl[_http_fdfs_dl.Test_Result == 'Completed'])
        if self.fdfs_dl_attempts > 0:
            self.fdfs_dl_ratio_failed = self.fdfs_dl_failed / self.fdfs_dl_attempts
            self.fdfs_dl_ratio_succes = self.fdfs_dl_succes / self.fdfs_dl_attempts
        else:
            self.fdfs_dl_ratio_failed = 0
            self.fdfs_dl_ratio_succes = 0

        try:
            self.fdfs_dl_ratio_cutoff = self.fdfs_dl_cutoff / (self.fdfs_dl_attempts - self.fdfs_dl_failed)
        except:
            self.fdfs_dl_ratio_cutoff = None

        # SMALL FILE UPLOAD
        _http_fdfs_ul = _records[(_records.Type_of_Test == 'httpTransfer') & (_records.Test_Name == 'FDFS http UL ST')]
        self.fdfs_ul_attempts = len(_http_fdfs_ul)
        self.fdfs_ul_failed   = len(_http_fdfs_ul[_http_fdfs_ul.Test_Result == 'Failed'])
        self.fdfs_ul_cutoff   = len(_http_fdfs_ul[_http_fdfs_ul.Test_Result == 'Cutoff'])
        self.fdfs_ul_succes   = len(_http_fdfs_ul[_http_fdfs_ul.Test_Result == 'Completed'])
        if self.fdfs_ul_attempts > 0:
            self.fdfs_ul_ratio_failed = self.fdfs_ul_failed / self.fdfs_ul_attempts
            self.fdfs_ul_ratio_succes = self.fdfs_ul_succes / self.fdfs_ul_attempts
        else:
            self.fdfs_ul_ratio_failed = 0
            self.fdfs_ul_ratio_succes = 0
        try:
            self.fdfs_ul_ratio_cutoff = self.fdfs_ul_cutoff / (self.fdfs_ul_attempts - self.fdfs_ul_failed)
        except:
            self.fdfs_ul_ratio_cutoff = None

        # CAPACITY DOWNLOAD
        _http_fdtt_dl = _records[(_records.Type_of_Test == 'httpTransfer') & (_records.Test_Name == 'FDTT http DL MT')]
        self.fdtt_dl_attempts = len(_http_fdtt_dl)
        self.fdtt_dl_failed   = len(_http_fdtt_dl[_http_fdtt_dl.Test_Result == 'Failed'])
        self.fdtt_dl_cutoff   = len(_http_fdtt_dl[_http_fdtt_dl.Test_Result == 'Cutoff'])
        self.fdtt_dl_succes   = len(_http_fdtt_dl[_http_fdtt_dl.Test_Result == 'Completed'])

        _tmp1 = _http_fdtt_dl["Mean_Data_Rate_Kbit_s"].describe(percentiles=_perc)
        self.fdtt_dl_min_mdr = _tmp1.loc['min']/1000
        self.fdtt_dl_p10_mdr = _tmp1.loc['10%']/1000
        self.fdtt_dl_avg_mdr = _tmp1.loc['mean']/1000
        self.fdtt_dl_p50_mdr = _tmp1.loc['50%']/1000
        self.fdtt_dl_p90_mdr = _tmp1.loc['90%']/1000
        self.fdtt_dl_max_mdr = _tmp1.loc['max']/1000

        # CAPACITY DOWNLOAD
        _http_fdtt_ul = _records[(_records.Type_of_Test == 'httpTransfer') & (_records.Test_Name == 'FDTT http UL MT')]
        self.0\
            ,= len(_http_fdtt_ul)
        self.fdtt_ul_failed   = len(_http_fdtt_ul[_http_fdtt_ul.Test_Result == 'Failed'])
        self.fdtt_ul_cutoff   = len(_http_fdtt_ul[_http_fdtt_ul.Test_Result == 'Cutoff'])
        self.fdtt_ul_succes   = len(_http_fdtt_ul[_http_fdtt_ul.Test_Result == 'Completed'])

        _tmp1 = _http_fdtt_ul["Mean_Data_Rate_Kbit_s"].describe(percentiles=_perc)
        self.fdtt_ul_min_mdr = _tmp1.loc['min']/1000
        self.fdtt_ul_p10_mdr = _tmp1.loc['10%']/1000
        self.fdtt_ul_avg_mdr = _tmp1.loc['mean']/1000
        self.fdtt_ul_p50_mdr = _tmp1.loc['50%']/1000
        self.fdtt_ul_p90_mdr = _tmp1.loc['90%']/1000
        self.fdtt_ul_max_mdr = _tmp1.loc['max']/1000

        try:
            _cntul = _http_tran[(_http_tran.Direction == 'UL') & (_http_tran.Test_Result != 'Failed')]
            _cntdl = _http_tran[(_http_tran.Direction == 'DL') & (_http_tran.Test_Result != 'Failed')]
            _tmpu1 = _http_tran[_http_tran.Direction == 'UL']
            _tmpd1 = _http_tran[_http_tran.Direction == 'DL']
            if len(_cntul) > 0:
                self.ul_qam = 1.0 * len( _tmpu1[_tmpu1.ShareUL64QAM > 0.3] ) / len(_cntul)
            else:
                self.ul_qam = 0
            if len(_cntdl) > 0:
                self.dl_qam = 1.0 * len( _tmpd1[_tmpd1.ShareDL256QAM > 0.1] ) / len(_cntdl)
            else:
                self.dl_qam = 0
        except:
            self.ul_qam = 0
            self.dl_qam = 0

        # BROWSING
        _http_brws_dl = _records[_records.Type_of_Test == 'httpBrowser']
        self.brws_dl_attempts = len(_http_brws_dl)
        self.brws_dl_failed   = len(_http_brws_dl[_http_brws_dl.Test_Result == 'Failed'])
        self.brws_dl_cutoff   = len(_http_brws_dl[_http_brws_dl.Test_Result == 'Cutoff'])
        self.brws_dl_succes   = len(_http_brws_dl[_http_brws_dl.Test_Result == 'Completed'])
        if self.brws_dl_attempts > 0:
            self.brws_dl_ratio_failed = self.brws_dl_failed / self.brws_dl_attempts
            self.brws_dl_ratio_succes = self.brws_dl_succes / self.brws_dl_attempts
        else:
            self.brws_dl_ratio_failed = 0
            self.brws_dl_ratio_succes = 0
        if (self.brws_dl_attempts - self.brws_dl_failed) > 0:
            self.brws_dl_ratio_cutoff = self.brws_dl_cutoff / (self.brws_dl_attempts - self.brws_dl_failed)
        else:
            self.brws_dl_ratio_cutoff = None
        _cmp = _http_brws_dl[_http_brws_dl.Test_Result == 'Completed']
        _tmp1 = _cmp["Transfer_Duration_s"].describe(percentiles=_perc)
        self.brws_dl_min_ctt = _tmp1.loc['min']
        self.brws_dl_p10_ctt = _tmp1.loc['10%']
        self.brws_dl_avg_ctt = _tmp1.loc['mean']
        self.brws_dl_p50_ctt = _tmp1.loc['50%']
        self.brws_dl_p90_ctt = _tmp1.loc['90%']
        self.brws_dl_max_ctt = _tmp1.loc['max']

        _tmp2 = _cmp["http_Browser_Access_Duration_s"].describe(percentiles=_perc)
        self.brws_dl_min_sat = _tmp2.loc['min']*1000
        self.brws_dl_p10_sat = _tmp2.loc['10%']*1000
        self.brws_dl_avg_sat = _tmp2.loc['mean']*1000
        self.brws_dl_p50_sat = _tmp2.loc['50%']*1000
        self.brws_dl_p90_sat = _tmp2.loc['90%']*1000
        self.brws_dl_max_sat = _tmp2.loc['max']*1000

        _tmp3 = _cmp["TCP_RTT_Service_Access_Delay_ms"].describe(percentiles=_perc)
        self.brws_dl_min_rtt = _tmp3.loc['min']
        self.brws_dl_p10_rtt = _tmp3.loc['10%']
        self.brws_dl_avg_rtt = _tmp3.loc['mean']
        self.brws_dl_p50_rtt = _tmp3.loc['50%']
        self.brws_dl_p90_rtt = _tmp3.loc['90%']
        self.brws_dl_max_rtt = _tmp3.loc['max']

        # VIDEO STREAM
        _vs_yt_dl = _records[_records.Type_of_Test == 'VideoStreaming']
        self.vs_yt_dl_attempts = len(_vs_yt_dl)
        self.vs_yt_dl_failed = len(_vs_yt_dl[_vs_yt_dl.Test_Result == 'Failed'])
        self.vs_yt_dl_cutoff = len(_vs_yt_dl[_vs_yt_dl.Test_Result == 'Cutoff'])
        self.vs_yt_dl_succes = len(_vs_yt_dl[_vs_yt_dl.Test_Result == 'Completed'])
        self.vs_yt_dl_irritating = len(_vs_yt_dl[_vs_yt_dl.Irritating_Video_Playout == 'True'])
        if self.vs_yt_dl_attempts > 0:
            self.vs_yt_dl_ratio_failed = self.vs_yt_dl_failed / self.vs_yt_dl_attempts
            self.vs_yt_dl_ratio_succes = self.vs_yt_dl_succes / self.vs_yt_dl_attempts
        else:
            self.vs_yt_dl_ratio_failed = 0
            self.vs_yt_dl_ratio_succes = 0
        if (self.vs_yt_dl_attempts - self.vs_yt_dl_failed) > 0:
            self.vs_yt_dl_ratio_cutoff = self.vs_yt_dl_cutoff / (self.vs_yt_dl_attempts - self.vs_yt_dl_failed)
        else:
            self.vs_yt_dl_ratio_cutoff = None
        if self.vs_yt_dl_succes > 0:
            self.vs_yt_dl_ratio_irritating = self.vs_yt_dl_irritating / self.vs_yt_dl_succes
        else:
            self.vs_yt_dl_ratio_irritating = None

        _vs_yt_dl = _vs_yt_dl[_vs_yt_dl.Test_Result == 'Completed']
        _tmp1 = _vs_yt_dl["VideoStream_VQ_Mean"].describe(percentiles=_perc)
        self.vs_yt_dl_min_vMOS = _tmp1.loc['min']
        self.vs_yt_dl_p10_vMOS = _tmp1.loc['10%']
        self.vs_yt_dl_avg_vMOS = _tmp1.loc['mean']
        self.vs_yt_dl_p50_vMOS = _tmp1.loc['50%']
        self.vs_yt_dl_p90_vMOS = _tmp1.loc['90%']
        self.vs_yt_dl_max_vMOS = _tmp1.loc['max']
        self.vs_yt_dl_good_vMOS = len(_vs_yt_dl[_vs_yt_dl.VideoStream_VQ_Mean >= 1.6])
        self.vs_yt_dl_poor_vMOS = self.vs_yt_dl_succes - self.vs_yt_dl_good_vMOS
        if self.vs_yt_dl_succes > 0:
            self.vs_yt_dl_prat_vMOS = self.vs_yt_dl_poor_vMOS / self.vs_yt_dl_succes
        else:
            self.vs_yt_dl_prat_vMOS = 0

        _tmp2 = _vs_yt_dl["VideoStream_Time_to_First_Picture_s"].describe(percentiles=_perc)
        self.vs_yt_dl_min_ttfp = _tmp2.loc['min']
        self.vs_yt_dl_p10_ttfp = _tmp2.loc['10%']
        self.vs_yt_dl_avg_ttfp = _tmp2.loc['mean']
        self.vs_yt_dl_p50_ttfp = _tmp2.loc['50%']
        self.vs_yt_dl_p90_ttfp = _tmp2.loc['90%']
        self.vs_yt_dl_max_ttfp = _tmp2.loc['max']
        self.vs_yt_dl_poor_ttfp = len(_vs_yt_dl[_vs_yt_dl.VideoStream_Time_to_First_Picture_s > 10.0])
        self.vs_yt_dl_good_ttfp = len(_vs_yt_dl[_vs_yt_dl.VideoStream_Time_to_First_Picture_s <= 10.0])
        if (self.vs_yt_dl_succes + self.vs_yt_dl_cutoff) > 0:
            self.vs_yt_dl_prat_ttfp = self.vs_yt_dl_poor_ttfp / (self.vs_yt_dl_succes + self.vs_yt_dl_cutoff)
        else:
            self.vs_yt_dl_prat_ttfp = 0

        # FACEBOOK UPLOAD
        _fb_fdfs_ul = _records[(_records.Type_of_Test == 'Application') & (_records.Test_Name == 'FDFS Facebook UL ST')]
        self.fb_fdfs_ul_attempts = len(_fb_fdfs_ul)
        self.fb_fdfs_ul_failed   = len(_fb_fdfs_ul[_fb_fdfs_ul.Test_Result == 'Failed'])
        self.fb_fdfs_ul_cutoff   = len(_fb_fdfs_ul[_fb_fdfs_ul.Test_Result == 'Cutoff'])
        self.fb_fdfs_ul_succes   = len(_fb_fdfs_ul[_fb_fdfs_ul.Test_Result == 'Completed'])
        if self.fb_fdfs_ul_attempts > 0:
            self.fb_fdfs_ul_ratio_failed = self.fb_fdfs_ul_failed / self.fb_fdfs_ul_attempts
            self.fb_fdfs_ul_ratio_succes = self.fb_fdfs_ul_succes / self.fb_fdfs_ul_attempts
        else:
            self.fb_fdfs_ul_ratio_failed = 0
            self.fb_fdfs_ul_ratio_succes = 0
        try:
            self.fb_fdfs_ul_ratio_cutoff = self.fb_fdfs_ul_cutoff / (self.fb_fdfs_ul_attempts - self.fb_fdfs_ul_failed)
        except:
            self.fb_fdfs_ul_ratio_cutoff = None

        _cmp = _fb_fdfs_ul[_fb_fdfs_ul.Test_Result == 'Completed']
        _tmp1 =_cmp["Transfer_Duration_s"].describe(percentiles=_perc)
        self.fb_fdfs_ul_min_ctt = _tmp1.loc['min']
        self.fb_fdfs_ul_p10_ctt = _tmp1.loc['10%']
        self.fb_fdfs_ul_avg_ctt = _tmp1.loc['mean']
        self.fb_fdfs_ul_p50_ctt = _tmp1.loc['50%']
        self.fb_fdfs_ul_p90_ctt = _tmp1.loc['90%']
        self.fb_fdfs_ul_max_ctt = _tmp1.loc['max']

        # WHATSAPP UPLOAD
        _wa_fdfs_ul = _records[(_records.Type_of_Test == 'Application') & (_records.Test_Name == 'FDFS WhatsApp UL')]
        self.wa_fdfs_ul_attempts = len(_wa_fdfs_ul)
        self.wa_fdfs_ul_failed = len(_wa_fdfs_ul[_wa_fdfs_ul.Test_Result == 'Failed'])
        self.wa_fdfs_ul_cutoff = len(_wa_fdfs_ul[_wa_fdfs_ul.Test_Result == 'Cutoff'])
        self.wa_fdfs_ul_succes = len(_wa_fdfs_ul[_wa_fdfs_ul.Test_Result == 'Completed'])
        if self.wa_fdfs_ul_attempts > 0:
            self.wa_fdfs_ul_ratio_failed = self.wa_fdfs_ul_failed / self.wa_fdfs_ul_attempts
            self.wa_fdfs_ul_ratio_succes = self.wa_fdfs_ul_succes / self.wa_fdfs_ul_attempts
        else:
            self.wa_fdfs_ul_ratio_failed = 0
            self.wa_fdfs_ul_ratio_succes = 0
        if (self.wa_fdfs_ul_attempts - self.wa_fdfs_ul_failed) > 0:
            self.wa_fdfs_ul_ratio_cutoff = self.wa_fdfs_ul_cutoff / (self.wa_fdfs_ul_attempts - self.wa_fdfs_ul_failed)
        else:
            self.wa_fdfs_ul_ratio_cutoff = None

        _cmp = _wa_fdfs_ul[_wa_fdfs_ul.Test_Result == 'Completed']
        _tmp1 = _cmp["Transfer_Duration_s"].describe(percentiles=_perc)
        self.wa_fdfs_ul_min_ctt = _tmp1.loc['min']
        self.wa_fdfs_ul_p10_ctt = _tmp1.loc['10%']
        self.wa_fdfs_ul_avg_ctt = _tmp1.loc['mean']
        self.wa_fdfs_ul_p50_ctt = _tmp1.loc['50%']
        self.wa_fdfs_ul_p90_ctt = _tmp1.loc['90%']
        self.wa_fdfs_ul_max_ctt = _tmp1.loc['max']

        # WHATSAPP DOWNLOAD
        _wa_fdfs_dl = _records[(_records.Type_of_Test == 'Application') & (_records.Test_Name == 'FDFS WhatsApp DL')]
        self.wa_fdfs_dl_attempts = len(_wa_fdfs_dl)
        self.wa_fdfs_dl_failed = len(_wa_fdfs_dl[_wa_fdfs_dl.Test_Result == 'Failed'])
        self.wa_fdfs_dl_cutoff = len(_wa_fdfs_dl[_wa_fdfs_dl.Test_Result == 'Cutoff'])
        self.wa_fdfs_dl_succes = len(_wa_fdfs_dl[_wa_fdfs_dl.Test_Result == 'Completed'])
        if self.wa_fdfs_dl_attempts > 0:
            self.wa_fdfs_dl_ratio_failed = self.wa_fdfs_dl_failed / self.wa_fdfs_dl_attempts
            self.wa_fdfs_dl_ratio_succes = self.wa_fdfs_dl_succes / self.wa_fdfs_dl_attempts
        else:
            self.wa_fdfs_dl_ratio_failed = 0
            self.wa_fdfs_dl_ratio_succes = 0
        if (self.wa_fdfs_dl_attempts - self.wa_fdfs_dl_failed) > 0:
            self.wa_fdfs_dl_ratio_cutoff = self.wa_fdfs_dl_cutoff / (self.wa_fdfs_dl_attempts - self.wa_fdfs_dl_failed)
        else:
            self.wa_fdfs_dl_ratio_cutoff = None
        _cmp = _wa_fdfs_dl[_wa_fdfs_dl.Test_Result == 'Completed']
        _tmp1 = _cmp["Transfer_Duration_s"].describe(percentiles=_perc)
        self.wa_fdfs_dl_min_ctt = _tmp1.loc['min']
        self.wa_fdfs_dl_p10_ctt = _tmp1.loc['10%']
        self.wa_fdfs_dl_avg_ctt = _tmp1.loc['mean']
        self.wa_fdfs_dl_p50_ctt = _tmp1.loc['50%']
        self.wa_fdfs_dl_p90_ctt = _tmp1.loc['90%']
        self.wa_fdfs_dl_max_ctt = _tmp1.loc['max']

# FUNCTIONS
########################################################################################################################
def append_kpis(glev, ind, ac, bc, cc):
    if cc.start_time is None:
        s_time = ac.start_time
    elif ac.start_time < cc.start_time:
        s_time = ac.start_time
    else:
        s_time = cc.start_time

    if ac.end_time is None:
        e_time = cc.end_time
    elif ac.end_time < cc.end_time:
        e_time = cc.end_time
    else:
        e_time = ac.end_time
    if e_time is None:
        e_time = s_time
    if s_time is None:
        s_time = e_time

    data = pd.DataFrame({"kpi_levels": glev,
                  "grp_index": ind,
                  "G_LEVEL_1": ac.g1,
                  "G_LEVEL_2": ac.g2,
                  "G_LEVEL_3": ac.g3,
                  "G_LEVEL_4": ac.g4,
                  "G_LEVEL_5": ac.g5,
                  "TRAIN_INFO": ac.train + '--' + ac.wagon,
                  "VENDOR": ac.vendor,
                  "START_TIME": s_time,
                  "END_TIME": e_time,
                  "CLASSIC_ATTEMPTS": ac.classic_attempts,
                  "CLASSIC_COMPLETED": ac.classic_completed,
                  "CLASSIC_FAILED": ac.classic_failed,
                  "CLASSIC_DROPPED": ac.classic_dropped,
                  "CLASSIC_CSSR": ac.classic_cssr,
                  "CLASSIC_DCR": ac.classic_dcr,
                  "CLASSIC_CSR": ac.classic_ccr,
                  "CLASSIC_MIN_CST": ac.classic_min_cst,
                  "CLASSIC_AVG_CST": ac.classic_avg_cst,
                  "CLASSIC_MAX_CST": ac.classic_max_cst,
                  "CLASSIC_P10_CST": ac.classic_p10_cst,
                  "CLASSIC_P50_CST": ac.classic_p50_cst,
                  "CLASSIC_P90_CST": ac.classic_p90_cst,
                  "CLASSIC_BAD_CST": ac.classic_poorCST,
                  "CLASSIC_OKK_CST": ac.classic_goodCST,
                  "CLASSIC_BAD_CST_RATIO": ac.classic_poorCSTratio,
                  "VOLTE_CM": ac.volte_start,
                  "VOLTE_CM_RATIO": ac.volte_ratio,
                  "VOLTE_CM_END_RATIO": ac.volte_end_ratio,
                  "CSFB_CM": ac.csfb_ratio,
                  "WHATSAPP_ATTEMPTS": ac.whatsapp_attempts,
                  "WHATSAPP_COMPLETED": ac.whatsapp_completed,
                  "WHATSAPP_FAILED": ac.whatsapp_failed,
                  "WHATSAPP_DROPPED": ac.whatsapp_dropped,
                  "WHATSAPP_CSSR": ac.whatsapp_cssr,
                  "WHATSAPP_DCR": ac.whatsapp_dcr,
                  "WHATSAPP_CSR": ac.whatsapp_ccr,
                  "WHATSAPP_MIN_CST": ac.whatsapp_min_cst,
                  "WHATSAPP_AVG_CST": ac.whatsapp_avg_cst,
                  "WHATSAPP_MAX_CST": ac.whatsapp_max_cst,
                  "WHATSAPP_P10_CST": ac.whatsapp_p10_cst,
                  "WHATSAPP_P50_CST": ac.whatsapp_p50_cst,
                  "WHATSAPP_P90_CST": ac.whatsapp_p90_cst,
                  "WHATSAPP_BAD_CST": ac.whatsapp_poorCST,
                  "WHATSAPP_OKK_CST": ac.whatsapp_goodCST,
                  "WHATSAPP_BAD_CST_RATIO": ac.whatsapp_poorCSTratio,
                  "CLASSIC_POLQA_ATTEMPTS": bc.classic_attempts,
                  "CLASSIC_POLQA_BAD": bc.classic_poorPMOS,
                  "CLASSIC_POLQA_BAD_RATIO": bc.classic_poorRMOS,
                  "CLASSIC_POLQA_MIN_MOS": bc.classic_min_mos,
                  "CLASSIC_POLQA_AVG_MOS": bc.classic_avg_mos,
                  "CLASSIC_POLQA_MAX_MOS": bc.classic_max_mos,
                  "CLASSIC_POLQA_P10_MOS": bc.classic_p10_mos,
                  "CLASSIC_POLQA_P50_MOS": bc.classic_p50_mos,
                  "CLASSIC_POLQA_P90_MOS": bc.classic_p90_mos,
                  "CLASSIC_POLQA_VOLTE_MIN_MOS": bc.classic_min_vmos,
                  "CLASSIC_POLQA_VOLTE_AVG_MOS": bc.classic_avg_vmos,
                  "CLASSIC_POLQA_VOLTE_MAX_MOS": bc.classic_max_vmos,
                  "CLASSIC_POLQA_VOLTE_P10_MOS": bc.classic_p10_vmos,
                  "CLASSIC_POLQA_VOLTE_P50_MOS": bc.classic_p50_vmos,
                  "CLASSIC_POLQA_VOLTE_P90_MOS": bc.classic_p90_vmos,
                  "WHATSAPP_POLQA_ATTEMPTS": bc.whatsapp_attempts,
                  "WHATSAPP_POLQA_BAD": bc.whatsapp_poorPMOS,
                  "WHATSAPP_POLQA_BAD_RATIO": bc.whatsapp_poorRMOS,
                  "WHATSAPP_POLQA_MIN_MOS": bc.whatsapp_min_mos,
                  "WHATSAPP_POLQA_AVG_MOS": bc.whatsapp_avg_mos,
                  "WHATSAPP_POLQA_MAX_MOS": bc.whatsapp_max_mos,
                  "WHATSAPP_POLQA_P10_MOS": bc.whatsapp_p10_mos,
                  "WHATSAPP_POLQA_P50_MOS": bc.whatsapp_p50_mos,
                  "WHATSAPP_POLQA_P90_MOS": bc.whatsapp_p90_mos,
                  "HTTP_TRANSFER_LTE_RATIO": cc.lte_share,
                  "HTTP_TRANSFER_LTE_UL_64QAM_RATIO": cc.ul_qam,
                  "HTTP_TRANSFER_LTE_DL_2564QAMRATIO": cc.dl_qam,
                  "HTTP_TRANSFER_FDFS_DL_ATTEMPTS": cc.fdfs_dl_attempts,
                  "HTTP_TRANSFER_FDFS_DL_FAILED": cc.fdfs_dl_failed,
                  "HTTP_TRANSFER_FDFS_DL_CUTOFF": cc.fdfs_dl_cutoff,
                  "HTTP_TRANSFER_FDFS_DL_SUCCESS": cc.fdfs_dl_succes,
                  "HTTP_TRANSFER_FDFS_DL_FAILED_RATIO": cc.fdfs_dl_ratio_failed,
                  "HTTP_TRANSFER_FDFS_DL_CUTOFF_RATIO": cc.fdfs_dl_ratio_cutoff,
                  "HTTP_TRANSFER_FDFS_DL_SUCCESS_RATIO": cc.fdfs_dl_ratio_succes,
                  "HTTP_TRANSFER_FDFS_UL_ATTEMPTS": cc.fdfs_ul_attempts,
                  "HTTP_TRANSFER_FDFS_UL_FAILED": cc.fdfs_ul_failed,
                  "HTTP_TRANSFER_FDFS_UL_CUTOFF": cc.fdfs_ul_cutoff,
                  "HTTP_TRANSFER_FDFS_UL_SUCCESS": cc.fdfs_ul_succes,
                  "HTTP_TRANSFER_FDFS_UL_FAILED_RATIO": cc.fdfs_ul_ratio_failed,
                  "HTTP_TRANSFER_FDFS_UL_CUTOFF_RATIO": cc.fdfs_ul_ratio_cutoff,
                  "HTTP_TRANSFER_FDFS_UL_SUCCESS_RATIO": cc.fdfs_ul_ratio_succes,
                  "HTTP_TRANSFER_FDTT_DL_ATTEMPTS": cc.fdtt_dl_attempts,
                  "HTTP_TRANSFER_FDTT_DL_FAILED": cc.fdtt_dl_failed,
                  "HTTP_TRANSFER_FDTT_DL_CUTOFF": cc.fdtt_dl_cutoff,
                  "HTTP_TRANSFER_FDTT_DL_SUCCESS": cc.fdtt_dl_succes,
                  "HTTP_TRANSFER_FDTT_DL_MDR_MIN": cc.fdtt_dl_min_mdr,
                  "HTTP_TRANSFER_FDTT_DL_MDR_P10": cc.fdtt_dl_p10_mdr,
                  "HTTP_TRANSFER_FDTT_DL_MDR_AVG": cc.fdtt_dl_avg_mdr,
                  "HTTP_TRANSFER_FDTT_DL_MDR_P50": cc.fdtt_dl_p50_mdr,
                  "HTTP_TRANSFER_FDTT_DL_MDR_P90": cc.fdtt_dl_p90_mdr,
                  "HTTP_TRANSFER_FDTT_DL_MDR_MAX": cc.fdtt_dl_max_mdr,
                  "HTTP_TRANSFER_FDTT_UL_ATTEMPTS": cc.fdtt_ul_attempts,
                  "HTTP_TRANSFER_FDTT_UL_FAILED": cc.fdtt_ul_failed,
                  "HTTP_TRANSFER_FDTT_UL_CUTOFF": cc.fdtt_ul_cutoff,
                  "HTTP_TRANSFER_FDTT_UL_SUCCESS": cc.fdtt_ul_succes,
                  "HTTP_TRANSFER_FDTT_UL_MDR_MIN": cc.fdtt_ul_min_mdr,
                  "HTTP_TRANSFER_FDTT_UL_MDR_P10": cc.fdtt_ul_p10_mdr,
                  "HTTP_TRANSFER_FDTT_UL_MDR_AVG": cc.fdtt_ul_avg_mdr,
                  "HTTP_TRANSFER_FDTT_UL_MDR_P50": cc.fdtt_ul_p50_mdr,
                  "HTTP_TRANSFER_FDTT_UL_MDR_P90": cc.fdtt_ul_p90_mdr,
                  "HTTP_TRANSFER_FDTT_UL_MDR_MAX": cc.fdtt_ul_max_mdr,
                  "HTTP_BROWSING_ATTEMPTS": cc.brws_dl_attempts,
                  "HTTP_BROWSING_FAILED": cc.brws_dl_failed,
                  "HTTP_BROWSING_CUTOFF": cc.brws_dl_cutoff,
                  "HTTP_BROWSING_SUCCESS": cc.brws_dl_succes,
                  "HTTP_BROWSING_FAILED_RATIO": cc.brws_dl_ratio_failed,
                  "HTTP_BROWSING_CUTOFF_RATIO": cc.brws_dl_ratio_cutoff,
                  "HTTP_BROWSING_SUCCESS_RATIO": cc.brws_dl_ratio_succes,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_MIN": cc.brws_dl_min_rtt,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_P10": cc.brws_dl_p10_rtt,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_AVG": cc.brws_dl_avg_rtt,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_P50": cc.brws_dl_p50_rtt,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_P90": cc.brws_dl_p90_rtt,
                  "HTTP_BROWSING_ROUNDTRIP_TIME_MAX": cc.brws_dl_max_rtt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_MIN": cc.brws_dl_min_ctt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_P10": cc.brws_dl_p10_ctt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_AVG": cc.brws_dl_avg_ctt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_P50": cc.brws_dl_p50_ctt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_P90": cc.brws_dl_p90_ctt,
                  "HTTP_BROWSING_CONTENT_TRANSFER_TIME_MAX": cc.brws_dl_max_ctt,
                  "VIDEO_STREAM_ATTEMPTS": cc.vs_yt_dl_attempts,
                  "VIDEO_STREAM_FAILED": cc.vs_yt_dl_failed,
                  "VIDEO_STREAM_CUTOFF": cc.vs_yt_dl_cutoff,
                  "VIDEO_STREAM_SUCCESS": cc.vs_yt_dl_succes,
                  "VIDEO_STREAM_IRRITATING_PLAYOUT": cc.vs_yt_dl_irritating,
                  "VIDEO_STREAM_FAILED_RATIO": cc.vs_yt_dl_ratio_failed,
                  "VIDEO_STREAM_CUTOFF_RATIO": cc.vs_yt_dl_ratio_cutoff,
                  "VIDEO_STREAM_SUCCESS_RATIO": cc.vs_yt_dl_ratio_succes,
                  "VIDEO_STREAM_IRRITATING_PLAYOUT_RATIO" : cc.vs_yt_dl_ratio_irritating,
                  "VIDEO_STREAM_VMOS_MIN": cc.vs_yt_dl_min_vMOS,
                  "VIDEO_STREAM_VMOS_P10": cc.vs_yt_dl_p10_vMOS,
                  "VIDEO_STREAM_VMOS_AVG": cc.vs_yt_dl_avg_vMOS,
                  "VIDEO_STREAM_VMOS_P50": cc.vs_yt_dl_p50_vMOS,
                  "VIDEO_STREAM_VMOS_P90": cc.vs_yt_dl_p90_vMOS,
                  "VIDEO_STREAM_VMOS_MAX": cc.vs_yt_dl_max_vMOS,
                  "VIDEO_STREAM_VMOS_BAD": cc.vs_yt_dl_poor_vMOS,
                  "VIDEO_STREAM_VMOS_OKK": cc.vs_yt_dl_good_vMOS,
                  "VIDEO_STREAM_VMOS_BAD_RATIO": cc.vs_yt_dl_prat_vMOS,
                  "VIDEO_STREAM_TTFP_MIN": cc.vs_yt_dl_min_ttfp,
                  "VIDEO_STREAM_TTFP_P10": cc.vs_yt_dl_p10_ttfp,
                  "VIDEO_STREAM_TTFP_AVG": cc.vs_yt_dl_avg_ttfp,
                  "VIDEO_STREAM_TTFP_P50": cc.vs_yt_dl_p50_ttfp,
                  "VIDEO_STREAM_TTFP_P90": cc.vs_yt_dl_p90_ttfp,
                  "VIDEO_STREAM_TTFP_MAX": cc.vs_yt_dl_max_ttfp,
                  "VIDEO_STREAM_TTFP_BAD": cc.vs_yt_dl_poor_ttfp,
                  "VIDEO_STREAM_TTFP_OKK": cc.vs_yt_dl_good_ttfp,
                  "VIDEO_STREAM_TTFP_BAD_RATIO": cc.vs_yt_dl_prat_ttfp,
                  "FACEBOOK_FDFS_UL_ATTEMPTS": cc.fb_fdfs_ul_attempts,
                  "FACEBOOK_FDFS_UL_FAILED": cc.fb_fdfs_ul_failed,
                  "FACEBOOK_FDFS_UL_CUTOFF": cc.fb_fdfs_ul_cutoff,
                  "FACEBOOK_FDFS_UL_SUCCESS": cc.fb_fdfs_ul_succes,
                  "FACEBOOK_FDFS_UL_FAILED_RATIO": cc.fb_fdfs_ul_ratio_failed,
                  "FACEBOOK_FDFS_UL_CUTOFF_RATIO": cc.fb_fdfs_ul_ratio_cutoff,
                  "FACEBOOK_FDFS_UL_SUCCESS_RATIO": cc.fb_fdfs_ul_ratio_succes,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_MIN": cc.fb_fdfs_ul_min_ctt,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_P10": cc.fb_fdfs_ul_p10_ctt,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_AVG": cc.fb_fdfs_ul_avg_ctt,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_P50": cc.fb_fdfs_ul_p50_ctt,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_P90": cc.fb_fdfs_ul_p90_ctt,
                  "FACEBOOK_FDFS_UL_TRANSFER_TIME_MAX": cc.fb_fdfs_ul_max_ctt}, index = [0])
    return data

def write_pd(df_tmp,book,sheet_name):
    sheet_tmp = book.create_sheet(sheet_name)
    j = 1
    for v in list(df_tmp.columns.values):
        try:
            sheet_tmp.cell(row=1, column=j).value = v
            sheet_tmp.title = sheet_name
        except:
            print('Header Printing Failed')
            pass
        j += 1
    i = 1
    for value in list(df_tmp):
        j = 2
        for v in list(df_tmp[value]):
            try:
                sheet_tmp.cell(row=j, column=i).value = v
            except:
                print('Writting in row', i, ' and collumn ', j, 'FAILED!')
                pass
            j += 1
        i += 1

def buildKPIs(cols,cdr_v,cdr_s,cdr_d,m1,m2,m3,m4):
    # LEVEL G0
    df = pd.DataFrame(columns=cols)
    a = Voice(cdr_v, g1 = '', g2 = '', g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
    b = Speech(cdr_s, g1 = '', g2 = '', g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
    c = Data(cdr_d, g1 = '', g2 = '', g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
    d1 = append_kpis('g0', 0, a, b, c)
    df = df.append(d1)
    del d1
    # LEVEL G2
    i = 0
    for module in [m1,m2,m3,m4]:
        i = i + 1
        a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
        b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
        c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
        d1 = append_kpis('g2', i, a, b, c)
        df = df.append(d1)
        del d1

        # G_Level_2 + Vendor
        j = 0
        for sub in module.g2v:
            tvend = sub.split(';')[1]
            if tvend != "":
                j = j+1
                a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = tvend, reg = '', t_name = '', w_name = '')
                b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = tvend, reg = '', t_name = '', w_name = '')
                c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = '', g4 = '', g5 = '', vend = tvend, reg = '', t_name = '', w_name = '')
                d1 = append_kpis('g2v', i*10+j, a, b, c)
                df = df.append(d1)
                del d1

        # G_Level_3
        j = 0
        for sub in module.g3:
            if sub != "":
                j = j+1
                a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = sub, g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = sub, g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = sub, g4 = '', g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                d1 = append_kpis('g3', i*10+j, a, b, c)
                df = df.append(d1)
                del d1

        # G_Level_3 + Vendor
        j = 0
        for sub in module.g3v:
            if sub.split(';')[1] != "":
                j = j+1
                a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 = '', g5 = '', vend = sub.split(';')[1], reg = '', t_name = '', w_name = '')
                b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 = '', g5 = '', vend = sub.split(';')[1], reg = '', t_name = '', w_name = '')
                c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 = '', g5 = '', vend = sub.split(';')[1], reg = '', t_name = '', w_name = '')
                d1 = append_kpis('g3v', i*10+j, a, b, c)
                df = df.append(d1)
                del d1

        # G_Level_4
        j = 0
        if module.g2 != 'Train Route':
            for sub in module.g4:
                if sub.split(';')[1] != "":
                    j = j+1
                    a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 =  sub.split(';')[1], g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                    b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 =  sub.split(';')[1], g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                    c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = sub.split(';')[0], g4 =  sub.split(';')[1], g5 = '', vend = '', reg = '', t_name = '', w_name = '')
                    d1 = append_kpis('g4', i*100+j, a, b, c)
                    df = df.append(d1)
                    del d1
        else:
            for sub in module.tn:
                if sub.split(';')[1] != "":
                    j = j+1
                    a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = '', g4 =  sub.split(';')[0], g5 = '', vend = '', reg = '', t_name = sub.split(';')[1], w_name = sub.split(';')[2])
                    b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = '', g4 =  sub.split(';')[0], g5 = '', vend = '', reg = '', t_name = sub.split(';')[1], w_name = sub.split(';')[2])
                    c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = '', g4 =  sub.split(';')[0], g5 = '', vend = '', reg = '',  t_name = sub.split(';')[1],  w_name = sub.split(';')[2])
                    d1 = append_kpis('g4', i*100+j, a, b, c)
                    df = df.append(d1)
                    del d1
        # G_Level_5
        j = 0
        for sub in module.g5:
            if sub != "":
                j = j+1
                a = Voice(cdr_v, g1 = module.g1, g2 = module.g2, g3 = '', g4 = sub.split(';')[0], g5 = sub.split(';')[1], vend = '', reg = '', t_name = '', w_name = '')
                b = Speech(cdr_s, g1 = module.g1, g2 = module.g2, g3 = '', g4 = sub.split(';')[0], g5 = sub.split(';')[1], vend = '', reg = '', t_name = '', w_name = '')
                c = Data(cdr_d, g1 = module.g1, g2 = module.g2, g3 = '', g4 = sub.split(';')[0], g5 = sub.split(';')[1], vend = '', reg = '', t_name = '', w_name = '')
                d1 = append_kpis('g5', i*1000+j, a, b, c)
                df = df.append(d1)
                del d1
    return df

# MAIN PROGRAM
########################################################################################################################

########################################################################################################################
#                                                    REQUIRES USER INPUT                                               #
########################################################################################################################
cdr_views = []

# OPERATOR 1
op = OperatorCDRs(server       = "blndb11",
                  voice_db     = "DE_BM_Voice_1905",
                  data_db      = "DE_BM_Data_1905",
                  voice_table  = "vVoiceCDR2018_Operator1",
                  speech_table = "vSpeechCDR2018_Operator1",
                  data_table   = "vDataCDR2018_Operator1",
                  output_table = 'NEW_KPI_OPERATOR_1')
cdr_views.append(op)
del op
########################################################################################################################
#                                                    END USER INPUT                                                    #
########################################################################################################################

# EXTRACT ALL POSSIBLE GEOLOCATIONS INFORMATION
for op in cdr_views:
    print('CONNECTING...\n',
          '\nSERVER            : ',op.srv,
          '\nVOICE DB          : ',op.v_db,
          '\nDATA DB           : ',op.d_db,
          '\nVOICE CDR TABLE   : ',op.v_cdr,
          '\nSPEECH CDR TABLE  : ',op.s_cdr,
          '\nDATA CDR TABLE    : ',op.d_cdr)
    # DATABASE CONNECTION
    vdb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( op.srv, op.v_db ) )
    ddb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( op.srv, op.d_db ) )
    vcur = vdb.cursor()
    dcur = ddb.cursor()

    print("DB CONNECTED, STARTING GEOLOCATION EXTRACTION...")
    fv = open("IMPORT_VoiceCDR_G_LEVELS.txt", "r+")
    sql_command = fv.read() + op.v_cdr
    cdr_voice = pd.read_sql(sql_command, vdb)
    fv.close()

    fd = open("IMPORT_DataCDR_G_LEVELS.txt", "r+")
    sql_command = fd.read() + op.d_cdr
    cdr_data = pd.read_sql(sql_command, vdb)
    fd.close()

    # DRIVE CITY MODULES
    dc = GeoStuff("Drive", "City")
    dc.updateLocations(cdr_voice)
    dc.updateLocations(cdr_data)
    # DRIVE CONNECTING ROUTES MODULES
    dr = GeoStuff("Drive", "Connecting Roads")
    dr.updateLocations(cdr_voice)
    dr.updateLocations(cdr_data)
    # WALK CITY MODULES
    wc = GeoStuff("Walk", "City")
    wc.updateLocations(cdr_voice)  #
    wc.updateLocations(cdr_data)
    # WALK TRAIN MODULES
    wt = GeoStuff("Walk", "Train Route")
    wt.updateLocations(cdr_voice)
    wt.updateLocations(cdr_data)

    print('EXTRACTED GEOLOCATIONS')
    dc.printLocations()
    dr.printLocations()
    wc.printLocations()
    wt.printLocations()

    # CLOSE DB CONNECTIONS AND DELETE TMC CDRs
    vdb.close()
    ddb.close()
    del cdr_voice
    del cdr_data
    print('\n...GEOLOCATION INFORMATION EXTRACTION SUCCESS!')

# READ REQUIRED KPIS IN KPI REPORT
fd = open("IMPORT_kpiReport_COLUMNS.txt", "r+")
colums_kpiReport = fd.readlines()
# delete \n from end of line
count = 0
for line in colums_kpiReport:
    colums_kpiReport[count] = line[:-1]
    count += 1

for op in cdr_views:
    print('\nKPI CALCULATION...\n')
    print('\nCONNECTING...\n',
          '\nSERVER            : ',op.srv,
          '\nVOICE DB          : ',op.v_db,
          '\nDATA DB           : ',op.d_db,
          '\nVOICE CDR TABLE   : ',op.v_cdr,
          '\nSPEECH CDR TABLE  : ',op.s_cdr,
          '\nDATA CDR TABLE    : ',op.d_cdr)

    # DATABASE CONNECTION
    vdb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( op.srv, op.v_db ) )
    ddb = pyodbc.connect(r'Driver={SQL Server};Server=%s;Database=%s;Trusted_Connection=yes;' %( op.srv, op.d_db ) )
    vcur = vdb.cursor()
    dcur = ddb.cursor()

    print("DB CONNECTED, EXTRACTING CDRs...")
    cdr_voice  = pd.read_sql("SELECT * FROM " + op.v_cdr, vdb)
    cdr_speech = pd.read_sql("SELECT * FROM " + op.s_cdr, vdb)
    cdr_data   = pd.read_sql("SELECT * FROM " + op.d_cdr, ddb)
    vcur.close()
    dcur.close()
    vdb.close()
    ddb.close()

    print("CDRs EXTRACTED:...")
    print("Voice Records  :", len(cdr_voice))
    print("Speech Records :", len(cdr_speech))
    print("Data Records   :", len(cdr_data))

    df = buildKPIs(colums_kpiReport, cdr_voice, cdr_speech, cdr_data, dc, dr, wc, wt)
    del cdr_voice
    del cdr_speech
    del cdr_data

    # WRITE KPI REPORT TO DB
    engine = sqlalchemy.create_engine(f'mssql+pyodbc://{op.srv}/{op.v_db}?driver=SQL+Server+Native+Client+11.0')
    df.to_sql("TM_KPI_REPORT", con=engine, if_exists='replace')
    book = Workbook()
    write_pd(df, book, 'KPI REPORT')
    book.save('tm_tmp.xlsx')
print(colums_kpiReport)
print(type(colums_kpiReport))

